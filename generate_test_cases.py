import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from collections import Counter, defaultdict

wb = openpyxl.Workbook()
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
wrap = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ['S.No', 'Test Case ID', 'Module / Feature', 'Test Case Description', 'Test Steps', 'Expected Result', 'Priority', 'Status', 'Remarks']
col_widths = [8, 14, 24, 42, 42, 42, 12, 12, 25]

# [ID, Module, Description, Steps, Expected, Priority]
tc = [
    # ===== LOGIN & AUTH =====
    ['TC_AUTH_001', 'Login', 'Valid admin login', '1. Open app\n2. Enter admin@sls.com / admin123\n3. Click Sign In', 'Redirect to Dashboard. Token stored in localStorage.', 'High'],
    ['TC_AUTH_002', 'Login', 'Valid driver login', '1. Enter driver@sls.com / admin123\n2. Click Sign In', 'Redirect to /driver-portal', 'High'],
    ['TC_AUTH_003', 'Login', 'Invalid email format', '1. Enter invalid email (e.g. abc)\n2. Click Sign In', 'HTML5 validation error. No API call.', 'Medium'],
    ['TC_AUTH_004', 'Login', 'Wrong password', '1. Enter valid email + wrong password\n2. Click Sign In', 'Toast error: Login failed', 'High'],
    ['TC_AUTH_005', 'Login', 'Unregistered email', '1. Enter unregistered email\n2. Click Sign In', 'Toast error from API', 'Medium'],
    ['TC_AUTH_006', 'Login', 'Show/hide password toggle', '1. Enter password\n2. Click eye icon', 'Password toggles between visible and masked', 'Low'],
    ['TC_AUTH_007', 'Login', 'Login page UI check', '1. Open login page', 'Left: Hero image + truck animation. Right: Login form. SLS Fleet branding. Create account link.', 'Low'],
    ['TC_AUTH_008', 'Signup', 'Valid signup submission', '1. Click Create account\n2. Fill all fields\n3. Click Register', 'Success message. Redirect to success screen with Back to Login button.', 'High'],
    ['TC_AUTH_009', 'Signup', 'Passwords do not match', '1. Enter different passwords\n2. Click Register', 'Toast: Passwords do not match', 'High'],
    ['TC_AUTH_010', 'Signup', 'Password less than 6 characters', '1. Enter short password\n2. Click Register', 'Toast: Password must be at least 6 characters', 'Medium'],
    ['TC_AUTH_011', 'Signup', 'Duplicate email signup', '1. Enter already registered email\n2. Fill other fields\n3. Register', 'API error: Email already registered', 'Medium'],
    ['TC_AUTH_012', 'Signup', 'Missing required field', '1. Leave Name blank\n2. Click Register', 'HTML5 required field validation message', 'Medium'],
    ['TC_AUTH_013', 'Logout', 'Admin logout', '1. Login as admin\n2. Click Logout in header', 'localStorage cleared. Redirect to /login.', 'High'],
    ['TC_AUTH_014', 'Logout', 'Driver logout', '1. Login as driver\n2. Click Logout in driver portal header', 'localStorage cleared. Redirect to /login.', 'High'],

    # ===== DASHBOARD =====
    ['TC_DASH_001', 'Dashboard', 'Dashboard metrics load correctly', '1. Login as admin\n2. View dashboard', 'Stats cards: Total Vehicles, Active Drivers, Active Tenders, Pending Approvals', 'High'],
    ['TC_DASH_002', 'Dashboard', 'Recent alerts section', '1. View dashboard alerts area', 'Alert items with entity name, doc type, priority badges (HIGH/MEDIUM). Max 5 items.', 'Medium'],
    ['TC_DASH_003', 'Dashboard', 'Welcome message personalization', '1. Login and view dashboard', 'Header shows: Welcome back, {user.name}', 'Low'],

    # ===== VEHICLES =====
    ['TC_VEH_001', 'Vehicles', 'Load vehicle list page', '1. Navigate to Vehicles page', 'Table: Vehicle No, Owner, Make, Plant, Status, Actions. Stats cards show totals.', 'High'],
    ['TC_VEH_002', 'Vehicles', 'Search by vehicle number', '1. Type vehicle number in search box', 'Table filters to matching vehicles (partial match)', 'High'],
    ['TC_VEH_003', 'Vehicles', 'Search by owner/plant/make', '1. Type owner name or plant in search', 'Table filters matching results', 'Medium'],
    ['TC_VEH_004', 'Vehicles', 'Filter by status', '1. Select Active from status dropdown', 'Only active vehicles displayed', 'High'],
    ['TC_VEH_005', 'Vehicles', 'Filter by plant', '1. Select a plant from plant dropdown', 'Only vehicles at selected plant shown', 'Medium'],
    ['TC_VEH_006', 'Vehicles', 'Combined search + filters', '1. Type search text\n2. Select status\n3. Select plant', 'All 3 filters applied together correctly', 'Medium'],
    ['TC_VEH_007', 'Vehicles', 'View vehicle details modal', '1. Click View button on a vehicle row', 'Modal opens with full vehicle details and documents', 'High'],
    ['TC_VEH_008', 'Vehicles', 'Add Vehicle button visible (allowed role)', '1. Login as maker/admin\n2. Go to Vehicles', 'Add Vehicle button is visible', 'High'],
    ['TC_VEH_009', 'Vehicles', 'Add Vehicle button hidden (restricted role)', '1. Login as viewer\n2. Go to Vehicles', 'Add Vehicle button NOT visible', 'High'],
    ['TC_VEH_010', 'Vehicles - Add', 'Step 1: Save vehicle details', '1. Click Add Vehicle\n2. Fill vehicle_no, owner_name, make\n3. Click Save & Upload Documents', 'Vehicle created. Proceed to Step 2.', 'High'],
    ['TC_VEH_011', 'Vehicles - Add', 'Step 1: Missing required fields', '1. Leave vehicle_no blank\n2. Click Save', 'Toast error: Please fill in required vehicle details', 'High'],
    ['TC_VEH_012', 'Vehicles - Add', 'Step 2: Upload document file', '1. Drag PDF to RC dropzone\n2. Set expiry date', 'File selected, filename displayed. Expiry date saved.', 'High'],
    ['TC_VEH_013', 'Vehicles - Add', 'Step 2: Invalid file type', '1. Try to upload .docx file', 'File rejected (only PDF, JPG, PNG accepted)', 'Medium'],
    ['TC_VEH_014', 'Vehicles - Add', 'Step 2: File exceeds 25MB', '1. Try to upload file larger than 25MB', 'Toast error about file size limit', 'Medium'],
    ['TC_VEH_015', 'Vehicles - Add', 'Step 3: Review and Submit', '1. Complete Steps 1 & 2\n2. Review summary\n3. Click Done', 'Summary shows all entered data. Navigate to My Submissions.', 'High'],

    # ===== DRIVERS =====
    ['TC_DRV_001', 'Drivers', 'Load driver list page', '1. Navigate to Drivers page', 'Table: Emp ID, Name, Phone, DL No, Allocated Vehicle, Status. Stats cards.', 'High'],
    ['TC_DRV_002', 'Drivers', 'Search drivers', '1. Type name/emp_id/phone in search', 'Table filters to matching drivers', 'High'],
    ['TC_DRV_003', 'Drivers', 'Filter by status', '1. Select Active from dropdown', 'Only active drivers shown', 'Medium'],
    ['TC_DRV_004', 'Drivers', 'View driver detail modal', '1. Click View on a driver row', 'Modal with driver details and uploaded documents', 'High'],
    ['TC_DRV_005', 'Drivers - Add', 'Step 1: Save driver details', '1. Click Add Driver\n2. Fill name, emp_id, phone, dl_no\n3. Click Save', 'Driver created. Proceed to Step 2.', 'High'],
    ['TC_DRV_006', 'Drivers - Add', 'Step 1: Missing required fields', '1. Leave emp_id blank\n2. Click Save', 'Toast error about required fields', 'High'],
    ['TC_DRV_007', 'Drivers - Add', 'Step 2: Upload DL and Hazardous cert', '1. Upload DL file with expiry\n2. Upload Hazardous cert with expiry', 'Files uploaded. Filenames: {emp_id}_dl.ext, {emp_id}_hazardous.ext', 'High'],
    ['TC_DRV_008', 'Drivers - Add', 'Step 3: Review and Submit', '1. Review driver summary\n2. Click Done', 'Navigate to My Submissions with pending status', 'High'],

    # ===== PLANTS =====
    ['TC_PLT_001', 'Plants', 'Load plants list', '1. Navigate to Plants page', 'Grid of plant cards: name, type, location, contact, vehicle stats', 'High'],
    ['TC_PLT_002', 'Plants', 'Plant vehicle stats', '1. View a plant card', 'Vehicle counts (Total, Active) match backend data', 'Medium'],
    ['TC_PLT_003', 'Plants', 'Add Plant button visibility', '1. Login as admin\n2. Go to Plants', 'Add Plant button visible only for admin/superuser', 'Medium'],

    # ===== STOPPAGES =====
    ['TC_STP_001', 'Stoppages', 'Load stoppages with analytics', '1. Navigate to Stoppages page', 'Stats: Total Stoppages, Currently Stopped (red), Resumed, Top Reasons', 'High'],
    ['TC_STP_002', 'Stoppages', 'Filter All/Active/Resumed', '1. Click Active filter button', 'Only currently stopped vehicles shown', 'High'],
    ['TC_STP_003', 'Stoppages', 'Stoppage details display', '1. View a stoppage record', 'Vehicle No, Date, Reason, Plant, Days Stopped visible', 'Medium'],
    ['TC_STP_004', 'Stoppages', 'Add Stoppage (allowed role)', '1. Login as plant_incharge\n2. Click Add Stoppage', 'Form opens. Can submit new stoppage.', 'High'],

    # ===== TENDERS =====
    ['TC_TND_001', 'Tenders', 'Load tenders page', '1. Navigate to Tenders', 'Cards with tender info. Stats: Total, Active, Expiring Soon, Vehicles Assigned.', 'High'],
    ['TC_TND_002', 'Tenders', 'Search and filter tenders', '1. Search by name/number/client\n2. Filter by status', 'Filtered results shown correctly', 'High'],
    ['TC_TND_003', 'Tenders', 'Create tender (3-tab form)', '1. Click Add Tender\n2. Fill Basic Info\n3. Fill Financial\n4. Assign vehicles\n5. Save', 'Tender created. All 3 tabs data saved.', 'High'],
    ['TC_TND_004', 'Tenders', 'Edit tender', '1. Click Edit on a tender\n2. Modify fields\n3. Save', 'Changes saved. Toast success.', 'High'],
    ['TC_TND_005', 'Tenders', 'Delete tender (admin only)', '1. Login as admin\n2. Click Delete\n3. Confirm', 'Tender deleted. Confirmation dialog shown.', 'High'],
    ['TC_TND_006', 'Tenders', 'Vehicle assignment to tender', '1. Go to Vehicles tab\n2. Click Assign', 'Vehicle added as pill. Can remove with X.', 'Medium'],
    ['TC_TND_007', 'Tenders', 'Days remaining color coding', '1. View tender cards', 'Green > 30 days, Amber 1-30 days, Red expired', 'Medium'],

    # ===== APPROVALS =====
    ['TC_APR_001', 'Approvals', 'Checker views pending items', '1. Login as checker\n2. Go to Approvals', 'Pending items with comment box, Approve/Reject buttons', 'High'],
    ['TC_APR_002', 'Approvals', 'Checker approves submission', '1. Add optional comment\n2. Click Verify & Forward to Approver', 'Status: checked. Checker info + timestamp saved.', 'High'],
    ['TC_APR_003', 'Approvals', 'Checker rejects submission', '1. Click Return to Maker', 'Status: rejected. Item removed from checker queue.', 'High'],
    ['TC_APR_004', 'Approvals', 'Approver views checked items', '1. Login as approver\n2. Go to Approvals', 'Only checked status items shown', 'High'],
    ['TC_APR_005', 'Approvals', 'Approver final approval', '1. Click Final Approve & Publish', 'Status: approved. Entity published.', 'High'],
    ['TC_APR_006', 'Approvals', 'Admin adds comment', '1. Login as admin\n2. View pending item\n3. Add comment', 'Comment saved. Visible to checker/approver. No action buttons.', 'Medium'],
    ['TC_APR_007', 'Approvals', 'Download document from approval', '1. Click download icon on attached doc', 'File downloads/opens in new tab', 'Medium'],

    # ===== MY SUBMISSIONS =====
    ['TC_SUB_001', 'My Submissions', 'View all submissions', '1. Login as maker\n2. Go to My Submissions', 'All submitted items. Stats: Pending, Awaiting, Approved, Rejected.', 'High'],
    ['TC_SUB_002', 'My Submissions', 'Filter by status', '1. Select status from dropdown', 'Only matching submissions shown', 'Medium'],
    ['TC_SUB_003', 'My Submissions', 'Progress timeline display', '1. View a submission card', 'Timeline: Submitted > Checked > Approved. Current step highlighted.', 'Medium'],
    ['TC_SUB_004', 'My Submissions', 'Rejected submission display', '1. View a rejected item', 'Red X mark. Rejected label. Reviewer comments visible.', 'Medium'],

    # ===== ALERTS =====
    ['TC_ALT_001', 'Alerts', 'Load alerts center', '1. Navigate to Alerts page', 'Stats: Total, Critical (red), Warnings (amber). Alert list loaded.', 'High'],
    ['TC_ALT_002', 'Alerts', 'Critical alert display', '1. View alert for expired document', 'Red styling, CRITICAL badge, negative days remaining', 'High'],
    ['TC_ALT_003', 'Alerts', 'Warning alert display', '1. View alert for doc expiring in 1-30 days', 'Amber styling, WARNING badge, positive days count', 'High'],
    ['TC_ALT_004', 'Alerts', 'Search alerts', '1. Type entity/document name in search', 'Filtered alert results', 'Medium'],
    ['TC_ALT_005', 'Alerts', 'Filter by severity and type', '1. Select Critical from filter\n2. Select Vehicles', 'Only matching alerts shown', 'Medium'],

    # ===== EXPIRY CALENDAR =====
    ['TC_CAL_001', 'Expiry Calendar', 'Calendar loads correctly', '1. Navigate to Expiry Calendar', 'Calendar renders. Summary: Total, Expired, Expiring 30 Days, Valid.', 'High'],
    ['TC_CAL_002', 'Expiry Calendar', 'Day colors match status', '1. View calendar days', 'Red = expired, Amber = expiring soon, Green = valid', 'High'],
    ['TC_CAL_003', 'Expiry Calendar', 'Click day shows documents', '1. Click a day with dot indicator', 'List of documents expiring that day shown below calendar', 'High'],
    ['TC_CAL_004', 'Expiry Calendar', 'Navigate months', '1. Click next/previous month arrow', 'Calendar shows selected month. Documents update.', 'Medium'],

    # ===== REPORTS =====
    ['TC_REP_001', 'Reports', 'Load reports page', '1. Navigate to Reports', 'Stats cards: Fleet Summary, Document Status, Contract Report', 'High'],
    ['TC_REP_002', 'Reports', 'Report type dropdown', '1. Select different report types from dropdown', 'Options: Fleet Summary, Vehicle Report, Driver Report, Document Status', 'Medium'],

    # ===== USER MANAGEMENT =====
    ['TC_USR_001', 'Users', 'Load users list', '1. Login as admin\n2. Navigate to Users', 'Table: Name, Email, Phone, Role, Status', 'High'],
    ['TC_USR_002', 'Users', 'Search users', '1. Type name/email/role in search', 'Filtered user results', 'Medium'],
    ['TC_USR_003', 'Users', 'Add new user', '1. Click Add User\n2. Fill: Name, Email, Phone, Role, Password\n3. Save', 'User created with status=active. Toast success.', 'High'],
    ['TC_USR_004', 'Users', 'Edit user', '1. Click Edit\n2. Change name, phone, role\n3. Save', 'User updated. Email field disabled (cannot change).', 'High'],
    ['TC_USR_005', 'Users', 'Deactivate/Activate user', '1. Click status toggle', 'User status changes between active/inactive', 'High'],
    ['TC_USR_006', 'Users', 'Access restricted to admin/superuser', '1. Login as viewer\n2. Try to access /users', 'Redirect. Sidebar link not visible.', 'High'],

    # ===== SIGNUP REQUESTS =====
    ['TC_SREQ_001', 'Signup Requests', 'Load pending requests', '1. Login as admin\n2. Go to Signup Requests', 'Cards: name, email, phone, timestamp. Approve/Reject buttons.', 'High'],
    ['TC_SREQ_002', 'Signup Requests', 'Approve request with role assignment', '1. Click Approve\n2. Select role from dropdown\n3. Click Approve & Activate', 'User activated with role. Removed from list.', 'High'],
    ['TC_SREQ_003', 'Signup Requests', 'Reject request', '1. Click Reject\n2. Confirm', 'Request deleted. Toast confirmation.', 'High'],
    ['TC_SREQ_004', 'Signup Requests', 'Approve without selecting role', '1. Click Approve\n2. Do not select role', 'Button disabled or toast error', 'Medium'],
    ['TC_SREQ_005', 'Signup Requests', 'Empty requests list', '1. No pending requests exist', 'No Pending Requests message shown', 'Low'],

    # ===== PROFILE =====
    ['TC_PRF_001', 'Profile', 'Load profile page', '1. Click profile link in sidebar/header', 'Profile: Photo, Name, Email, Phone, Emp ID, Role, Status', 'High'],
    ['TC_PRF_002', 'Profile', 'Upload profile photo', '1. Click camera icon\n2. Select JPG/PNG file', 'Photo uploads and updates immediately', 'High'],
    ['TC_PRF_003', 'Profile', 'Invalid photo format', '1. Try to upload .gif or .bmp', 'Toast error: Please upload JPG, PNG or WEBP', 'Medium'],
    ['TC_PRF_004', 'Profile', 'Photo exceeds 10MB', '1. Try to upload large image', 'Toast error: Photo must be under 10MB', 'Medium'],
    ['TC_PRF_005', 'Profile', 'Edit profile and submit for approval', '1. Click Edit\n2. Change name/phone\n3. Click Submit for Approval', 'Edit submitted. Notification shows pending approval.', 'High'],

    # ===== DRIVER PORTAL - ACCESS & LAYOUT =====
    ['TC_DP_001', 'Driver Portal - Access', 'Driver login redirects to portal', '1. Login as driver@sls.com / admin123', 'Redirected to /driver-portal. Sidebar visible with 3 navigation items.', 'High'],
    ['TC_DP_002', 'Driver Portal - Access', 'Non-driver cannot access portal', '1. Login as admin\n2. Navigate to /driver-portal', 'Redirect to dashboard or access denied', 'High'],
    ['TC_DP_003', 'Driver Portal - Layout', 'White header bar matches admin portal', '1. Login as driver\n2. View header', 'White background header. Welcome message, SOS, Refresh, Logout buttons.', 'Medium'],
    ['TC_DP_004', 'Driver Portal - Layout', 'Gray sidebar with navigation items', '1. View sidebar', 'Gray (bg-gray-300) sidebar. Items: My Profile, My Vehicle Detail, Vehicle Document. SLS Fleet logo.', 'Medium'],
    ['TC_DP_005', 'Driver Portal - Layout', 'Grid background pattern visible', '1. View main content area', 'Grid pattern background (matching admin portal) visible behind content cards', 'Low'],

    # ===== DRIVER PORTAL - NAVIGATION =====
    ['TC_DP_006', 'Driver Portal - Navigation', 'Sidebar click switches section content', '1. Click My Profile\n2. Click My Vehicle Detail\n3. Click Vehicle Document', 'Content changes to show selected section only. Active item highlighted (dark bg, white text).', 'High'],
    ['TC_DP_007', 'Driver Portal - Navigation', 'Mobile hamburger menu', '1. Open on mobile viewport\n2. Click hamburger icon', 'Sidebar slides in with overlay. Click overlay or item closes sidebar.', 'Medium'],
    ['TC_DP_008', 'Driver Portal - Navigation', 'Driver info in sidebar footer', '1. View sidebar bottom area', 'Driver photo/icon, name, email, Driver role displayed', 'Low'],
    ['TC_DP_009', 'Driver Portal - Navigation', 'Refresh button reloads data', '1. Click Refresh button in header', 'API called again. Data refreshed without page reload.', 'Medium'],

    # ===== DRIVER PORTAL - MY PROFILE =====
    ['TC_DP_P_001', 'Driver Portal - My Profile', 'Profile header card displays', '1. Click My Profile in sidebar', 'Large photo/icon, driver name, Employee ID, Driver badge visible', 'High'],
    ['TC_DP_P_002', 'Driver Portal - My Profile', 'Personal information card', '1. View My Profile section', 'Full Name, Employee ID, Phone (with icon), DL Number (with icon) displayed', 'High'],
    ['TC_DP_P_003', 'Driver Portal - My Profile', 'DL status - Valid', '1. Driver has DL expiry > 30 days from today', 'Green background, VALID badge, expiry date shown', 'High'],
    ['TC_DP_P_004', 'Driver Portal - My Profile', 'DL status - Expiring Soon', '1. Driver has DL expiry within 30 days', 'Amber background, N days badge with count', 'High'],
    ['TC_DP_P_005', 'Driver Portal - My Profile', 'DL status - Expired', '1. Driver has DL expiry date in the past', 'Red background, EXPIRED badge', 'High'],
    ['TC_DP_P_006', 'Driver Portal - My Profile', 'Hazardous certification status', '1. View hazardous cert section', 'Correct color badge (green/amber/red) with expiry date', 'High'],

    # ===== DRIVER PORTAL - MY VEHICLE DETAIL =====
    ['TC_DP_V_001', 'Driver Portal - Vehicle Detail', 'GPS map shows as first element', '1. Click My Vehicle Detail in sidebar', 'Vehicle Location GPS map card is the FIRST element on the page', 'High'],
    ['TC_DP_V_002', 'Driver Portal - Vehicle Detail', 'Vehicle stats cards display', '1. View vehicle section', 'Two stat cards: Vehicle No (blue icon), Current Plant (green icon)', 'Medium'],
    ['TC_DP_V_003', 'Driver Portal - Vehicle Detail', 'Vehicle details card', '1. View vehicle details card', 'All fields: Vehicle Number, Make, Owner, Vehicle Type, Chassis No, Engine No', 'High'],
    ['TC_DP_V_004', 'Driver Portal - Vehicle Detail', 'Current assignment card', '1. Vehicle has plant assigned', 'Blue gradient card showing Plant Location and Contract/Tender name', 'Medium'],
    ['TC_DP_V_005', 'Driver Portal - Vehicle Detail', 'No vehicle assigned message', '1. Driver has no allocated vehicle', 'No Vehicle Assigned message with Contact your supervisor text', 'High'],

    # ===== DRIVER PORTAL - VEHICLE DOCUMENT =====
    ['TC_DP_D_001', 'Driver Portal - Vehicle Doc', 'Documents list displays', '1. Click Vehicle Document in sidebar', 'All vehicle docs shown: RC, Insurance, Fitness, Tax, PUC, Permit, National Permit', 'High'],
    ['TC_DP_D_002', 'Driver Portal - Vehicle Doc', 'Document status - Valid', '1. View document with expiry > 30 days', 'Green background, VALID badge, expiry date', 'High'],
    ['TC_DP_D_003', 'Driver Portal - Vehicle Doc', 'Document status - Expiring Soon', '1. View document expiring within 30 days', 'Amber background, Nd badge (days remaining)', 'High'],
    ['TC_DP_D_004', 'Driver Portal - Vehicle Doc', 'Document status - Expired', '1. View expired document', 'Red background, EXPIRED badge', 'High'],
    ['TC_DP_D_005', 'Driver Portal - Vehicle Doc', 'Document count summary', '1. View bottom of documents card', 'Total Documents count. Warning if any documents need renewal.', 'Medium'],
    ['TC_DP_D_006', 'Driver Portal - Vehicle Doc', 'Download/view uploaded document', '1. Click View on an uploaded document', 'File opens in new tab or downloads', 'High'],
    ['TC_DP_D_007', 'Driver Portal - Vehicle Doc', 'No documents message', '1. Vehicle has no documents', 'No documents uploaded for this vehicle yet message', 'Medium'],

    # ===== DRIVER PORTAL - SOS =====
    ['TC_SOS_001', 'Driver Portal - SOS', 'Open SOS modal', '1. Click red SOS button in header', 'Modal: Emergency SOS title (red), type selection grid, message textarea', 'High'],
    ['TC_SOS_002', 'Driver Portal - SOS', 'Select SOS type', '1. Click Breakdown type button', 'Button highlighted with red border/background. Others deselected.', 'High'],
    ['TC_SOS_003', 'Driver Portal - SOS', 'Send SOS without selecting type', '1. Click Send SOS Alert without selecting type', 'Toast error: Please select an emergency type', 'High'],
    ['TC_SOS_004', 'Driver Portal - SOS', 'Send SOS successfully', '1. Select type (e.g. Accident)\n2. Add optional message\n3. Click Send SOS Alert', 'Toast: SOS alert sent successfully! Modal closes. Includes vehicle_no, driver_name.', 'High'],
    ['TC_SOS_005', 'Driver Portal - SOS', 'Cancel SOS modal', '1. Open SOS modal\n2. Click Cancel', 'Modal closes. No alert sent. Form reset.', 'Medium'],
    ['TC_SOS_006', 'Driver Portal - SOS', 'SOS sending loading state', '1. Click Send SOS Alert', 'Button shows Sending... text and is disabled during API call', 'Low'],

    # ===== RBAC =====
    ['TC_RBAC_001', 'RBAC', 'Driver restricted to portal only', '1. Login as driver', 'Only /driver-portal accessible. No admin sidebar or pages.', 'High'],
    ['TC_RBAC_002', 'RBAC', 'Viewer cannot add vehicle/driver', '1. Login as viewer\n2. Go to Vehicles/Drivers', 'Add buttons NOT visible', 'High'],
    ['TC_RBAC_003', 'RBAC', 'Checker sees Approvals page', '1. Login as checker\n2. Check sidebar', 'Approvals link visible and page accessible', 'High'],
    ['TC_RBAC_004', 'RBAC', 'Maker sees My Submissions', '1. Login as maker\n2. Check sidebar', 'My Submissions link visible and functional', 'High'],
    ['TC_RBAC_005', 'RBAC', 'Only admin/superuser access Users', '1. Login as checker\n2. Check sidebar', 'Users link NOT visible. /users redirects.', 'High'],
    ['TC_RBAC_006', 'RBAC', 'Sidebar shows role-based items', '1. Login with different roles\n2. Compare sidebar', 'Each role sees only permitted navigation items', 'High'],

    # ===== NAVIGATION =====
    ['TC_NAV_001', 'Navigation', 'Admin sidebar links work', '1. Click each sidebar link', 'Navigates to correct page for each menu item', 'High'],
    ['TC_NAV_002', 'Navigation', 'Active link highlighting', '1. Navigate to Vehicles page', 'Vehicles link highlighted (dark bg, white text). Others normal.', 'Medium'],
    ['TC_NAV_003', 'Navigation', 'Mobile responsive sidebar', '1. Resize to mobile\n2. Click hamburger menu', 'Sidebar slides in with overlay. Click overlay closes.', 'Medium'],
    ['TC_NAV_004', 'Navigation', 'Header notification bell', '1. Click bell icon in header', 'Navigate to /alerts page', 'Medium'],
    ['TC_NAV_005', 'Navigation', 'Header profile link', '1. Click profile avatar/name in header', 'Navigate to /profile page', 'Medium'],

    # ===== INTEGRATION =====
    ['TC_INT_001', 'Integration', 'Full vehicle approval workflow', '1. Maker creates vehicle\n2. Checker approves\n3. Approver final approves', 'Vehicle: pending > checked > approved. Visible in My Submissions.', 'High'],
    ['TC_INT_002', 'Integration', 'Document expiry triggers alert', '1. Vehicle has doc expiring in 10 days', 'Alert in Alerts page. Calendar marks day amber.', 'High'],
    ['TC_INT_003', 'Integration', 'Driver sees allocated vehicle in portal', '1. Admin allocates vehicle to driver\n2. Driver logs in', 'Driver portal shows vehicle details, documents, GPS.', 'High'],
    ['TC_INT_004', 'Integration', 'New user signup to login flow', '1. User signs up\n2. Admin approves with role\n3. User logs in', 'User can access role-specific features after approval.', 'High'],

    # ===== ERROR HANDLING =====
    ['TC_ERR_001', 'Error Handling', 'API server down', '1. Stop backend server\n2. Try to load a page', 'Error toast displayed. UI does not crash.', 'High'],
    ['TC_ERR_002', 'Error Handling', 'Session expired (401)', '1. Let JWT token expire\n2. Try any API action', 'Redirect to /login page automatically', 'High'],
    ['TC_ERR_003', 'Error Handling', 'Loading states visible', '1. Navigate to any data page', 'Spinner visible during API calls. Buttons disabled during processing.', 'Medium'],

    # ===== UI/UX =====
    ['TC_UX_001', 'UI/UX', 'Mobile responsive layout', '1. Open app on mobile viewport', 'Forms stack vertically. Sidebar hidden. Hamburger menu. Cards full width.', 'Medium'],
    ['TC_UX_002', 'UI/UX', 'Desktop layout', '1. Open app on desktop', 'Sidebar always visible. 2-column layouts. Proper spacing.', 'Medium'],
    ['TC_UX_003', 'UI/UX', 'Modal close with ESC key', '1. Open any modal\n2. Press ESC', 'Modal closes', 'Low'],
    ['TC_UX_004', 'UI/UX', 'Grid background pattern', '1. View admin and driver portal', 'Subtle grid lines visible on background (60px spacing)', 'Low'],
]

# === SHEET 1: Test Cases ===
ws = wb.active
ws.title = 'Test Cases'
for i, w in enumerate(col_widths):
    ws.column_dimensions[chr(65 + i)].width = w

# Title
ws.merge_cells('A1:I1')
ws['A1'].value = 'SLS Fleet Management System - Test Cases'
ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

# Subtitle
ws.merge_cells('A2:I2')
ws['A2'].value = f'Total Test Cases: {len(tc)} | Date: 2026-02-20 | Application: SLS Fleet Management System'
ws['A2'].font = Font(name='Calibri', size=10, color='666666')
ws['A2'].alignment = Alignment(horizontal='center')

# Headers row 4
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

ws.freeze_panes = 'A5'
ws.auto_filter.ref = f'A4:I{4 + len(tc)}'

# Write data
for idx, row_data in enumerate(tc):
    row = idx + 5
    ws.cell(row=row, column=1, value=idx + 1).alignment = center
    ws.cell(row=row, column=2, value=row_data[0]).alignment = center
    ws.cell(row=row, column=3, value=row_data[1]).alignment = wrap
    ws.cell(row=row, column=4, value=row_data[2]).alignment = wrap
    ws.cell(row=row, column=5, value=row_data[3]).alignment = wrap
    ws.cell(row=row, column=6, value=row_data[4]).alignment = wrap

    # Priority coloring
    p = ws.cell(row=row, column=7, value=row_data[5])
    p.alignment = center
    if row_data[5] == 'High':
        p.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        p.font = Font(color='9C0006', bold=True)
    elif row_data[5] == 'Medium':
        p.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        p.font = Font(color='9C6500', bold=True)
    else:
        p.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        p.font = Font(color='006100', bold=True)

    ws.cell(row=row, column=8, value='').alignment = center
    ws.cell(row=row, column=9, value='').alignment = wrap

    # Borders
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = thin_border

    # Highlight Driver Portal rows with light green
    if 'Driver Portal' in row_data[1]:
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if col != 7:  # dont override priority color
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    ws.row_dimensions[row].height = 52

# Status dropdown validation
dv = DataValidation(type='list', formula1='"Pass,Fail,Not Tested,Blocked,Skipped"', allow_blank=True)
dv.error = 'Select: Pass, Fail, Not Tested, Blocked, or Skipped'
dv.errorTitle = 'Invalid Status'
ws.add_data_validation(dv)
for r in range(5, 5 + len(tc)):
    dv.add(ws.cell(row=r, column=8))

# === SHEET 2: Summary ===
ws2 = wb.create_sheet('Summary')
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 15

ws2.merge_cells('A1:C1')
ws2['A1'].value = 'Test Execution Summary'
ws2['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws2['A1'].alignment = Alignment(horizontal='center')

for col, h in enumerate(['Module', 'Count', 'Priority H/M/L'], 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

mc = Counter()
mp = defaultdict(lambda: [0, 0, 0])
for t in tc:
    m = t[1].split(' - ')[0]
    mc[m] += 1
    i = 0 if t[5] == 'High' else 1 if t[5] == 'Medium' else 2
    mp[m][i] += 1

r = 4
for m in dict.fromkeys([t[1].split(' - ')[0] for t in tc]):
    ws2.cell(row=r, column=1, value=m).border = thin_border
    ws2.cell(row=r, column=2, value=mc[m]).border = thin_border
    ws2.cell(row=r, column=2).alignment = center
    p = mp[m]
    ws2.cell(row=r, column=3, value=f'{p[0]}/{p[1]}/{p[2]}').border = thin_border
    ws2.cell(row=r, column=3).alignment = center
    r += 1

# Total row
ws2.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
ws2.cell(row=r, column=1).border = thin_border
ws2.cell(row=r, column=2, value=len(tc)).font = Font(bold=True)
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).alignment = center
h_count = sum(1 for t in tc if t[5] == 'High')
m_count = sum(1 for t in tc if t[5] == 'Medium')
l_count = sum(1 for t in tc if t[5] == 'Low')
ws2.cell(row=r, column=3, value=f'{h_count}/{m_count}/{l_count}').font = Font(bold=True)
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).alignment = center

# Save
output_path = r'C:\Users\Archana\Downloads\SLS_Fleet_Test_Cases.xlsx'
wb.save(output_path)
print(f'Excel saved: {output_path}')
print(f'Total: {len(tc)} test cases (High={h_count}, Medium={m_count}, Low={l_count})')
